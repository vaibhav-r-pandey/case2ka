from openpyxl import load_workbook
import html2text
import gpt4ifx
from gpt4ifx import Gpt4ifxAPI

import warnings
warnings.simplefilter("ignore")

def writeCaseStringToJson(json_case, case_number):
    file_name = f"{case_number}.json"
    with open(file_name, 'w') as f:
        f.write(json_case)


def ExcelToJson(Case_Number, Case_Subject, Case_Description, Case_Description_file):
    print("ExcelToJson")
    case_string = []

    case_string.append({"Case Subject" : Case_Subject})
    case_string.append({"Case Description" : Case_Description})
    # Open up the Excel file.
    case_table = load_workbook(Case_Description_file).active
    
    #To access the table1 we need to activate and store to an another object (Here it is table2)
    for row in range(2, case_table.max_row):
        case_string.append({"Comment Number" : case_table[row][8].value})
        case_string.append({"Description" : html2text.html2text(case_table[row][3].value)})
        case_string.append({"Date Created" : case_table[row][5].value})
        case_string.append({"From" : case_table[row][6].value})  

    json_case = json.dumps(case_string, indent=4, sort_keys=True, default=str)
    writeCaseStringToJson(json_case, Case_Number)


def Generatesummary_GPT4IFX(json_case, case_number):    
    llm = gpt4ifx.GPT4IFX()
    # llm.Gpt4ifx_get_Bearertoken()
    llm_response = llm.Gpt4ifx_Chat(json_case)
    if llm_response != None:    
        with open(f"{case_number}.md", 'w') as f:
            f.write(llm_response)
        return llm_response
    else:
        return None
    

def Generatekba_GPT4IFX(json_case, case_number):
    gpt4ifx_api = Gpt4ifxAPI()
    llm_response = gpt4ifx_api.Gpt4ifx_kba(json_case)
    print(llm_response)
    if llm_response is not None:
        with open(f"{case_number}.md", 'w') as f:
            f.write(llm_response)
        return llm_response
    else:
        return None

def MSDReadCaseData(Case_Number, Case_Subject, Case_Description, Case_Description_file):
    case_string = []

    ExcelToJson(Case_Number, Case_Subject, Case_Description, Case_Description_file)
    with open(f"{Case_Number}.json", 'r') as f:
        json_case = f.read()

    kba_responce = Generatekba_GPT4IFX(json_case, Case_Number)

    return kba_responce





